"""Excel report rendering for the gravimetric CV measurement.

Separated from ``cv_mass_measurement.py`` so the measurement orchestration
stays free of spreadsheet-formatting detail. This module owns the result
workbook: its heat-map styling config and the writer. The measurement run
passes its own metadata (tip, syringe, temperature, density) in — those
stay configured in ``cv_mass_measurement.py``; edit the colours here.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import PatternFill

if TYPE_CHECKING:
    from cv_mass_measurement import VolumeResult

log = logging.getLogger("cv_mass_measurement")

# --- Result heat-map gradient ---------------------------------------------
# Each quality metric (systematic error, relative error, SD, CV) is shaded
# by the magnitude of its value across a per-metric band:
#     magnitude <= *_GREEN  -> fully green  (within acceptance / on target)
#     magnitude >= *_RED    -> fully red    (out of acceptance)
#     in between            -> green -> yellow -> red
# Basis = absolute value of the figure shown in the cell (sign is ignored;
# the number itself carries over/under direction). Fixed limits keep a
# given colour meaning the same across runs and tips, so tips stay
# comparable. Set *_GREEN to 0 for a plain 0-to-*_RED ramp; widen *_GREEN
# to give each metric an acceptance deadband that stays fully green.
# Accuracy (vs target):
SYS_ERR_GREEN_UL = 0.0  # |systematic error uL| at/below -> green
SYS_ERR_RED_UL = 1.0  # |systematic error uL| at/above -> red
REL_ERR_GREEN_PCT = 0.0  # |relative error %| at/below -> green
REL_ERR_RED_PCT = 5.0  # |relative error %| at/above -> red
# Precision (replicate scatter):
SD_GREEN_UL = 0.0  # SD uL at/below -> green
SD_RED_UL = 0.5  # SD uL at/above -> red
CV_GREEN_PCT = 0.0  # CV % (=100*SD/mean) at/below -> green
CV_RED_PCT = 5.0  # CV % at/above -> red

# Where the colour passes through yellow within each green->red band, as a
# fraction 0..1 (0.5 = yellow at the band midpoint; lower = yellow sooner).
HEATMAP_MID_FRACTION = 0.5

# Fill colours (RGB tuples, and ARGB/RGB hex for the blue). Light blue tags
# the header row and the target-volume column; the three heat-map stops are
# Excel's conventional good/neutral/bad pastels. Edit to restyle.
FILL_LIGHT_BLUE = "DDEBF7"
HEATMAP_GOOD = (198, 239, 206)  # light green  (#C6EFCE)
HEATMAP_MID = (255, 235, 156)  # light yellow (#FFEB9C)
HEATMAP_BAD = (255, 199, 206)  # light red    (#FFC7CE)


def _solid_fill(rgb_hex: str) -> PatternFill:
    """Return a solid cell fill for an ARGB/RGB hex string (no ``#``)."""
    return PatternFill(start_color=rgb_hex, end_color=rgb_hex, fill_type="solid")


def _heatmap_fill(
    magnitude: float, green_limit: float, red_limit: float
) -> PatternFill:
    """Green→yellow→red fill for ``magnitude`` over a green/red band.

    ``magnitude`` is the absolute value shown in the cell. At or below
    ``green_limit`` it is fully green (on target / within acceptance); at
    or above ``red_limit`` it is fully red. In between it ramps through
    yellow, positioned by ``HEATMAP_MID_FRACTION``. Used to shade the four
    quality columns (systematic error, relative error, SD, CV).

    Args:
        magnitude: Non-negative figure of merit for the cell.
        green_limit: Magnitude at or below which the colour is fully green.
        red_limit: Magnitude at or above which the colour is fully red.

    Returns:
        A solid :class:`PatternFill` at the interpolated colour.
    """
    span = red_limit - green_limit
    if span <= 0:
        fraction = 0.0 if magnitude <= green_limit else 1.0
    else:
        fraction = min(max((magnitude - green_limit) / span, 0.0), 1.0)
    mid = HEATMAP_MID_FRACTION
    if fraction <= mid:
        start, end = HEATMAP_GOOD, HEATMAP_MID
        local = fraction / mid if mid > 0 else 1.0
    else:
        start, end = HEATMAP_MID, HEATMAP_BAD
        local = (fraction - mid) / (1.0 - mid) if mid < 1 else 1.0
    channels = (round(start[i] + (end[i] - start[i]) * local) for i in range(3))
    return _solid_fill("".join(f"{c:02X}" for c in channels))


def write_workbook(
    results: list[VolumeResult],
    path: Path,
    *,
    tip_gauge: str,
    syringe_uL: int,
    lab_temp_c: float,
    water_density_g_per_ml: float,
) -> None:
    """Write the per-volume replicate masses and statistics to ``path``.

    The result table is colour-coded: the header row and target-volume
    column get a light-blue fill, and the CV / RSD / SD cells get a
    green→red heat-map by magnitude (see the gradient config constants).

    Args:
        results: One :class:`VolumeResult` per target volume, in order.
        path: Destination .xlsx file.
        tip_gauge: Syringe tip gauge (G), recorded in the metadata block.
        syringe_uL: Syringe volume (uL), recorded in the metadata block.
        lab_temp_c: Lab temperature (C), recorded in the metadata block.
        water_density_g_per_ml: Water density used for mass→volume.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CV Results"
    blue_fill = _solid_fill(FILL_LIGHT_BLUE)

    # Replicate counts can differ per volume, so size the Trial columns to
    # the widest row and pad the rest.
    max_replicates = max((len(r.masses_g) for r in results), default=0)

    # Run metadata block.
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meta = [
        ("Measured at", stamp),
        ("Syringe tip gauge (G)", tip_gauge),
        ("Syringe (uL)", syringe_uL),
        ("Lab temperature (C)", lab_temp_c),
        ("Water density (g/mL)", water_density_g_per_ml),
        (
            "Systematic error (uL)",
            "mean-target — accuracy, signed (+over/-under)",
        ),
        ("Relative error (%)", "100*(mean-target)/target — accuracy, signed"),
        ("SD (uL)", "replicate standard deviation — precision"),
        ("CV (%)", "100*SD/mean — coefficient of variation (= RSD)"),
    ]
    for label, value in meta:
        sheet.append([label, value])
    sheet.append([])

    header = (
        ["Target Volume (uL)"]
        + [f"Trial {i} (g)" for i in range(1, max_replicates + 1)]
        + [
            "Mean Volume (uL)",
            "Systematic error (uL)",
            "Relative error (%)",
            "SD (uL)",
            "CV (%)",
        ]
    )
    sheet.append(header)
    # Column layout: target, trials, mean, then the four quality metrics.
    sys_err_col = 1 + max_replicates + 2
    rel_err_col = sys_err_col + 1
    sd_col = sys_err_col + 2
    cv_col = sys_err_col + 3
    header_row = sheet.max_row
    for column in range(1, cv_col + 1):
        sheet.cell(row=header_row, column=column).fill = blue_fill

    for result in results:
        row = [result.target_uL]
        row += [round(mass, 4) for mass in result.masses_g]
        # Pad shorter rows so the summary columns stay aligned.
        row += [None] * (max_replicates - len(result.masses_g))
        row += [
            round(result.mean_volume_uL, 3),
            round(result.sys_error_uL, 3),
            round(result.rel_error_pct, 2),
            round(result.sd_uL, 3),
            round(result.cv_pct, 2),
        ]
        sheet.append(row)
        data_row = sheet.max_row
        sheet.cell(row=data_row, column=1).fill = blue_fill
        sheet.cell(row=data_row, column=sys_err_col).fill = _heatmap_fill(
            abs(result.sys_error_uL), SYS_ERR_GREEN_UL, SYS_ERR_RED_UL
        )
        sheet.cell(row=data_row, column=rel_err_col).fill = _heatmap_fill(
            abs(result.rel_error_pct), REL_ERR_GREEN_PCT, REL_ERR_RED_PCT
        )
        sheet.cell(row=data_row, column=sd_col).fill = _heatmap_fill(
            result.sd_uL, SD_GREEN_UL, SD_RED_UL
        )
        sheet.cell(row=data_row, column=cv_col).fill = _heatmap_fill(
            result.cv_pct, CV_GREEN_PCT, CV_RED_PCT
        )

    workbook.save(path)
    log.info("Wrote results to %s", path)
